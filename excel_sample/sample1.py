'''
Excel

install:
  pip install openpyxl
  pip install pandas
  pip install xlrd
'''
import collections
import openpyxl
from openpyxl.chart import BarChart, Reference
import pandas as pd

''' ファイル読み込み '''
excel_bookname = 'Book1.xlsx'
df_master = pd.read_excel(excel_bookname, sheet_name='マスターデータ')
df_data = pd.read_excel(excel_bookname, sheet_name='集計するデータ')

''' ファイル書き込み '''
sheet_name = '集計結果'
book = openpyxl.load_workbook(excel_bookname)
del book[sheet_name]
book.create_sheet(title=sheet_name)
sheet = book[sheet_name]
row_number = 1
chartPosition = 2
for master_column in df_master.columns:
    sheet.cell(row=row_number, column=1).value = master_column
    row_number += 1
    chartStartRow = row_number

    data = collections.Counter(sum(df_data[master_column].str.split(','), []))
    for key in df_master[master_column]:
        if str(key) == 'nan':
            break
        sheet.cell(row=row_number, column=1).value = key

        if data[key]:
            sheet.cell(row=row_number, column=2).value = data[key]
        else:
            sheet.cell(row=row_number, column=2).value = 0

        row_number += 1
    chartEndRow = row_number - 1
    row_number += 1

    # グラフ生成
    chart = BarChart()
    chart.title = master_column
    chart.legend = None

    titles = Reference(
        sheet, min_col=1, min_row=chartStartRow,
        max_row=chartEndRow)
    values = Reference(
        sheet, min_col=2, min_row=chartStartRow,
        max_col=2, max_row=chartEndRow)
    chart.add_data(values)
    chart.set_categories(titles)

    sheet.add_chart(chart, "E{}".format(chartPosition))
    chartPosition += 15

book.save(excel_bookname)
