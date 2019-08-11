'''
install:
  pip install openpyxl
  pip install pandas
  pip install xlrd
'''
import collections
import openpyxl
from openpyxl.chart import BarChart, Reference
import pandas as pd


excel_bookname = 'ブック名.xlsx'
data_sheet_name = '集計データ'
result_sheet_name = '集計結果'
target_column = '集計カラム'

# 集計
df_data = pd.read_excel(excel_bookname, sheet_name=data_sheet_name)
result_data = collections.Counter(
    sum(df_data[target_column].str.split(','), []))

# 書き込み、集計結果シートは作り直す
book = openpyxl.load_workbook(excel_bookname)
del book[result_sheet_name]
book.create_sheet(title=result_sheet_name)
sheet = book[result_sheet_name]

row_number = 0
for key in result_data:
    row_number += 1
    sheet.cell(row=row_number, column=1).value = key
    sheet.cell(row=row_number, column=2).value = result_data[key]

# 棒グラフ、凡例なし
chart = BarChart()
chart.title = '棒グラフタイトル'
chart.legend = None
values = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=row_number)
titles = Reference(sheet, min_col=1, min_row=1, max_row=row_number)
chart.add_data(values)
chart.set_categories(titles)
sheet.add_chart(chart, "D2")

book.save(excel_bookname)
